import requests
from time import sleep
from bs4 import BeautifulSoup
from openpyxl import workbook, load_workbook
#path = 'C:\Program Files (x86)\chromedriver.exe'
URL = "https://www.coingecko.com/"
#URL = 'https://www.neuralnine.com'
res = requests.get(URL).text
sleep(5)
soup = BeautifulSoup(res,'lxml')

wb = load_workbook('crypto.xlsx')
ws = wb.active
print(ws['A2'].value)

name = soup.find('table')
thead = name.find('thead')
tbody = name.find('tbody')
trr=tbody.find_all('tr')
#print(trr)
for j in trr:
    pagenumber=j.find(class_='table-number')
    coin=j.find(class_='lg:tw-flex')
    symbol=j.find(class_='d-lg-inline')
    volum=j.find(class_='td-liquidity_score')
    print(pagenumber.text,coin.text,symbol.text,volum.text)
    ws.append([pagenumber.text, coin.text, symbol.text, volum.text])
wb.save('crypto.xlsx')
#print(trr)




#td = tbody.find_all('td')
#th = thead.find_all('th')
"""for i in td:
    #th = tr.find('th')
    print( i.text)
#th = name.find_all('th')
#for i in th:
    #print(i.text)
    """